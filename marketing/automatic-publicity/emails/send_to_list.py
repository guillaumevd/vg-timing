import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import load_workbook


def send_email(name, recipient_email, sender_email, password):
    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = recipient_email
    message["Subject"] = "Système de chronométrage basé sur la technologie RFID active"

    email_body = f"""
    Cher(e) {name},

    Je me permets de vous contacter car je suis actuellement étudiant et développe un projet innovant de système de chronométrage basé sur la technologie RFID active. Mon objectif est de proposer une solution performante et abordable pour les clubs de modélisme, les pistes de karting, les événements de rallye, les clubs de course de vélo et autres organisations similaires.

    Avant de me lancer dans la réalisation de ce projet, je souhaiterais évaluer l'intérêt du marché pour un tel système. Contrairement aux systèmes de chronométrage plus coûteux comme ceux de MYLAPS (environ 3000 €) ou Race Result (autour de 2000 €), mon système de chronométrage serait proposé à un prix beaucoup plus abordable, autour de 500 €. De plus, je tiens à vous assurer que la précision et la fiabilité de mon système seront comparables à celles des solutions plus onéreuses mentionnées ci-dessus.

    Je suis convaincu que cette solution pourrait apporter une réelle valeur ajoutée à votre club ou organisation en améliorant la précision et la fiabilité du chronométrage de vos compétitions et entraînements, tout en offrant un excellent rapport qualité-prix.

    Afin de mieux comprendre les besoins du marché, je vous serais très reconnaissant si vous pouviez me faire part de votre opinion sur ce projet et de votre intérêt potentiel pour un tel système de chronométrage. Si vous êtes intéressé(e) ou si vous souhaitez en savoir plus sur mon projet, n'hésitez pas à me contacter. Je me tiens à votre disposition pour discuter de cette opportunité et répondre à vos questions.

    Je vous remercie par avance pour votre attention et votre réponse.

    Cordialement,

    Guillaume Vandriessche
    """

    message.attach(MIMEText(email_body, "plain"))

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(sender_email, recipient_email, message.as_string())


def main():
    sender_email = "Guillaumevandriessche776@gmail.com"
    password = "tmcdwhsltelmqdxa"

    excel_file = "contacts.xlsx"
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, recipient_email = row

        if name and recipient_email:
            try:
                send_email(name, recipient_email, sender_email, password)
                print(f"Email sent to {name} at {recipient_email}")
            except Exception as e:
                print(f"Error sending email to {name} at {recipient_email}: {e}")
        else:
            print(f"Invalid data in row: {row}")


if __name__ == "__main__":
    main()
